import ast
import os
import sys
import time
import json
import re
import requests
import subprocess
import pandas as pd
from datetime import datetime
from dotenv import load_dotenv
import shutil

# Load environment variables (e.g., OPENROUTER_API_KEY)
load_dotenv()

OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")

if not OPENROUTER_API_KEY:
    print("WARNING: OPENROUTER_API_KEY not found in environment or .env file.")
    print("Please add it to the .env file.")

# Model fallback chain: try DeepSeek free models first, then other free models, then direct Gemini
OPENROUTER_MODELS = [
    "deepseek/deepseek-chat:free",         # DeepSeek V3 — fast, strong at code
    "deepseek/deepseek-r1:free",           # DeepSeek R1 — reasoning model
    "tngtech/deepseek-r1t-chimera:free",   # R1+V3 merged
    "stepfun/step-3.5-flash:free",         # StepFun fallback
    "google/gemini-2.5-flash-lite",        # Gemini via OpenRouter
]
# Direct Gemini API as last-resort fallback
GEMINI_DIRECT_MODEL = "gemini-2.0-flash"

LOG_FILE = "experiment_logs.csv"
EXCEL_FILE = "experiment_logs.xlsx"
MAX_DRAWDOWN_THRESHOLD = 0.08  # Must match program.md constraint
BEST_MODEL_DIR = "best_model"  # Permanent home for the current champion

def query_gemini_direct(prompt, system_prompt=""):
    """Call Google Gemini API directly as a last-resort fallback."""
    if not GEMINI_API_KEY:
        print("GEMINI_API_KEY not set. Cannot use direct Gemini fallback.")
        return None, None

    url = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_DIRECT_MODEL}:generateContent?key={GEMINI_API_KEY}"
    payload = {"contents": [{"parts": [{"text": prompt}]}]}
    if system_prompt:
        payload["systemInstruction"] = {"parts": [{"text": system_prompt}]}

    try:
        response = requests.post(url, json=payload, timeout=60)
        response.raise_for_status()
        result = response.json()
        return result['candidates'][0]['content']['parts'][0]['text'], f"gemini-direct/{GEMINI_DIRECT_MODEL}"
    except Exception as e:
        print(f"Error querying Gemini direct API: {e}")
        return None, None

def query_llm(prompt, system_prompt=""):
    """Try each model in the fallback chain. Returns (response_text, model_used) tuple."""
    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "HTTP-Referer": "https://github.com/karpathy/autoresearch",
        "X-Title": "Autoresearch Quant Agent"
    }

    for model in OPENROUTER_MODELS:
        print(f"  Trying {model}...")
        payload = {
            "model": model,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": prompt}
            ]
        }
        try:
            response = requests.post("https://openrouter.ai/api/v1/chat/completions", headers=headers, json=payload, timeout=120)
            response.raise_for_status()
            result = response.json()
            text = result['choices'][0]['message']['content']
            if text:
                print(f"  Success with {model}")
                return text, model
        except Exception as e:
            print(f"  Failed {model}: {e}")
            continue

    # All OpenRouter models failed — try direct Gemini API
    print("All OpenRouter models failed. Trying direct Gemini API...")
    return query_gemini_direct(prompt, system_prompt)

def extract_code(text):
    """Extract python code from markdown code blocks."""
    match = re.search(r"```python\n(.*?)\n```", text, re.DOTALL)
    if match:
        return match.group(1)
    
    # If no python tag, try generic code block
    match = re.search(r"```\n(.*?)\n```", text, re.DOTALL)
    if match:
        return match.group(1)
        
    return text  # Assume the whole response is code if no blocks found

def save_excel_report():
    """Save a formatted Excel report with all experiment results and highlight the best model."""
    if not os.path.exists(LOG_FILE):
        return

    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed. Skipping Excel report.")
        return

    logs = pd.read_csv(LOG_FILE)
    if logs.empty:
        return

    # Find the best valid run
    valid = logs[(logs['Status'] == 'Completed') & (logs['val_max_drawdown'] <= MAX_DRAWDOWN_THRESHOLD)]
    best_iter = valid.loc[valid['val_excess_return'].idxmax(), 'Iteration'] if not valid.empty else None

    # Mark the best run
    logs['Is_Best'] = logs['Iteration'] == best_iter

    # Write to Excel with formatting
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
        logs.to_excel(writer, sheet_name='Experiment Log', index=False)
        ws = writer.sheets['Experiment Log']

        # Styling
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        best_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        constraint_fail_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Format headers
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Format data rows
        for row_idx in range(2, ws.max_row + 1):
            status_cell = ws.cell(row=row_idx, column=4)  # Status column
            is_best_cell = ws.cell(row=row_idx, column=ws.max_column)  # Is_Best column

            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

            # Highlight best row green
            if is_best_cell.value == True:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = best_fill
            # Highlight failed rows red
            elif status_cell.value and 'Failed' in str(status_cell.value):
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fail_fill
            # Highlight completed but over drawdown limit yellow
            elif status_cell.value == 'Completed':
                dd_cell = ws.cell(row=row_idx, column=6)  # val_max_drawdown column
                try:
                    if dd_cell.value and float(dd_cell.value) > MAX_DRAWDOWN_THRESHOLD:
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = constraint_fail_fill
                except (ValueError, TypeError):
                    pass

        # Auto-fit column widths
        for col_idx in range(1, ws.max_column + 1):
            max_len = len(str(ws.cell(row=1, column=col_idx).value or ''))
            for row_idx in range(2, min(ws.max_row + 1, 50)):  # Sample first 50 rows
                val = ws.cell(row=row_idx, column=col_idx).value
                max_len = max(max_len, len(str(val or '')))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

        # Add summary sheet
        summary_data = {
            'Metric': [
                'Total Iterations',
                'Completed Runs',
                'Failed Runs',
                'Timeout Runs',
                'Runs Meeting Drawdown Constraint',
                'Best Excess Return (Valid)',
                'Best Drawdown (Valid)',
                'Best Iteration #',
                'Best Model Used',
                'Drawdown Threshold',
            ],
            'Value': [
                len(logs),
                len(logs[logs['Status'] == 'Completed']),
                len(logs[logs['Status'].str.contains('Failed', na=False)]),
                len(logs[logs['Status'] == 'Timeout']),
                len(valid),
                valid['val_excess_return'].max() if not valid.empty else 'N/A',
                valid.loc[valid['val_excess_return'].idxmax(), 'val_max_drawdown'] if not valid.empty else 'N/A',
                int(best_iter) if best_iter is not None else 'N/A',
                valid.loc[valid['val_excess_return'].idxmax(), 'Model'] if not valid.empty else 'N/A',
                MAX_DRAWDOWN_THRESHOLD,
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

        # Format summary sheet
        ws2 = writer.sheets['Summary']
        for col_idx in range(1, 3):
            cell = ws2.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        ws2.column_dimensions['A'].width = 35
        ws2.column_dimensions['B'].width = 30
        for row_idx in range(2, ws2.max_row + 1):
            for col_idx in range(1, 3):
                ws2.cell(row=row_idx, column=col_idx).border = thin_border

    print(f"Excel report saved to {EXCEL_FILE}")

def run_experiment(iteration_num):
    print(f"\n[{datetime.now().strftime('%H:%M:%S')}] Starting iteration {iteration_num}")
    
    # 1. Read the current context
    with open("program.md", "r") as f:
         program_rules = f.read()
         
    with open("train.py", "r") as f:
         current_train_code = f.read()
         
    # Read history to give context — last 20 runs so the LLM doesn't repeat failed approaches
    history = ""
    if os.path.exists(LOG_FILE):
        logs_df = pd.read_csv(LOG_FILE)
        history_df = logs_df.tail(20)
        history = "\nRecent Experiment History (last 20 runs):\n" + history_df.to_string()
        # Also surface the single best run so the LLM knows the current champion
        valid_logs = logs_df[(logs_df['Status'] == 'Completed') & (logs_df['val_max_drawdown'] <= MAX_DRAWDOWN_THRESHOLD)]
        if not valid_logs.empty:
            best_row = valid_logs.loc[valid_logs['val_excess_return'].idxmax()]
            history += f"\n\nCurrent Best Run (champion to beat):\n{best_row.to_string()}"
         
    # 2. Formulate Prompt
    system_prompt = """You are a world-class Quantitative AI Researcher with deep expertise in machine learning,
feature engineering, and algorithmic trading. You have mastery of PyTorch (including Transformers, attention,
residual networks, and custom architectures), scikit-learn, and advanced ML techniques.
You think creatively and are not afraid to make bold architectural changes."""
    prompt = f"""
    Here are the overarching rules for your algorithmic trading portfolio optimizations:
    {program_rules}

    {history}

    Here is the current code for `train.py`:
    ```python
    {current_train_code}
    ```

    ## Your Task
    Rewrite `train.py` to maximize `val_excess_return` while keeping `val_max_drawdown` below {MAX_DRAWDOWN_THRESHOLD} ({MAX_DRAWDOWN_THRESHOLD*100:.0f}% max drawdown).

    You have FULL FREEDOM to make any changes you want. You can and should:
    - **Change the ML algorithm entirely**: Replace the neural network with Random Forest,
      Gradient Boosting, SVR, or any sklearn ensemble. You can also combine multiple models.
    - **Build deep learning architectures**: Use PyTorch to build Transformers, residual networks,
      attention-based models, Mixture of Experts, or any custom neural architecture.
    - **Engineer new features**: Create interaction features, polynomial features, rolling statistics,
      cross-sectional ranks, z-scores, or any derived signals from the raw tensor data.
      The features are already standardized — you can create new ones from them.
    - **Change the loss function**: Use ranking losses, Sharpe-based losses, custom objectives, or
      any loss that better aligns with the portfolio selection task.
    - **Change the training strategy**: Use cross-validation, early stopping, learning rate schedules,
      gradient clipping, or any training technique.
    - **Change the portfolio construction**: Modify how predictions are converted to weights —
      use softmax weighting, risk parity, volatility targeting, or any allocation scheme.
    - **Add market regime detection**: Build a regime classifier that goes to cash during downturns.
    - **Make MULTIPLE changes at once** if they complement each other.

    The only libraries guaranteed available are: torch, pandas, numpy, sklearn (scikit-learn).
    Do NOT use xgboost or lightgbm unless you include a try/except fallback to sklearn.

    ## Output Format
    1. First, briefly explain your hypothesis and what changes you're making (2-4 sentences).
    2. Then return the ENTIRE completely valid python script for `train.py` inside a single ```python codeblock.
       Do not omit any code or use comments like "# rest of code here". Write the full modified file.
    """
    
    # 3. Get LLM Proposal
    print("Querying LLM for new hypothesis...")
    llm_response, model_used = query_llm(prompt, system_prompt)

    if not llm_response:
        print("Failed to get response from all LLMs. Skipping iteration.")
        return False

    print(f"Using model: {model_used}")
        
    new_code = extract_code(llm_response)

    # Validate syntax before touching train.py
    try:
        ast.parse(new_code)
    except SyntaxError as e:
        print(f"Syntax error in generated code: {e}. Skipping iteration.")
        return False

    # Backup current best train.py
    shutil.copyfile("train.py", "train.py.best")

    # Write new experiment
    with open("train.py", "w") as f:
        f.write(new_code)
        
    # 4. Run the backtest
    print("Running training simulation (May take up to 5 mins)...")
    try:
        # Run with timeout matching our internal py timeout (+ buffer)
        result = subprocess.run([sys.executable, "train.py"], capture_output=True, text=True, timeout=330)
        output = result.stdout
        errors = result.stderr
        
        # 5. Parse KPIs
        val_excess_return = None
        val_max_drawdown = None
        
        # Look for [FINAL RESULTS] block outputs defined in train.py
        excess_match = re.search(r"val_excess_return=([0-9.\-]+)", output)
        if excess_match: val_excess_return = float(excess_match.group(1))
            
        drawdown_match = re.search(r"val_max_drawdown=([0-9.\-]+)", output)
        if drawdown_match: val_max_drawdown = float(drawdown_match.group(1))
        
        status = "Completed" if (val_excess_return is not None and val_max_drawdown is not None) else "Failed_Parse"
        if result.returncode != 0:
            status = "Failed_Run"
            print(f"Error during runtime:\n{errors}")
            
    except subprocess.TimeoutExpired:
        status = "Timeout"
        output = "Process exceeded 5 minute + buffer timeout."
        val_excess_return, val_max_drawdown = None, None
        
    # 6. Logging
    log_entry = {
         "Iteration": iteration_num,
         "Timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
         "Model": model_used,
         "Status": status,
         "val_excess_return": val_excess_return,
         "val_max_drawdown": val_max_drawdown,
         "Meets_Constraint": status == "Completed" and val_max_drawdown is not None and val_max_drawdown <= MAX_DRAWDOWN_THRESHOLD,
         "Hypothesis_Summary": llm_response.split("```")[0].strip()[:200].replace('\n', ' ')
    }

    df = pd.DataFrame([log_entry])
    if os.path.exists(LOG_FILE):
        df.to_csv(LOG_FILE, mode='a', header=False, index=False)
    else:
        df.to_csv(LOG_FILE, index=False)

    # Also write formatted Excel file with all results
    save_excel_report()

    print(f"Iteration Results: Status={status} | Excess Return={val_excess_return} | Drawdown={val_max_drawdown}")

    # 7. Rollback or Keep (Greedy Search)
    best_excess = float('-inf')
    if os.path.exists(LOG_FILE):
        logs = pd.read_csv(LOG_FILE)
        valid_logs = logs[(logs['Status'] == 'Completed') & (logs['val_max_drawdown'] <= MAX_DRAWDOWN_THRESHOLD)]
        if not valid_logs.empty:
            best_excess = valid_logs['val_excess_return'].max()

    is_new_best = (status == "Completed" and val_excess_return is not None
                   and val_excess_return > best_excess
                   and val_max_drawdown is not None
                   and val_max_drawdown <= MAX_DRAWDOWN_THRESHOLD)

    if is_new_best:
         print("NEW BEST MODEL FOUND! Accepting changes and archiving!")
         # Archive to iteration-specific directory
         archive_dir = f"saved_models/iteration_{iteration_num}"
         os.makedirs(archive_dir, exist_ok=True)
         shutil.copyfile("train.py", os.path.join(archive_dir, "train.py"))
         if os.path.exists("best_model.pt"):
             shutil.copyfile("best_model.pt", os.path.join(archive_dir, "best_model.pt"))
         # Always keep the current champion in best_model/ for easy access
         os.makedirs(BEST_MODEL_DIR, exist_ok=True)
         shutil.copyfile("train.py", os.path.join(BEST_MODEL_DIR, "train.py"))
         if os.path.exists("best_model.pt"):
             shutil.copyfile("best_model.pt", os.path.join(BEST_MODEL_DIR, "best_model.pt"))
         # Write metadata about the champion
         with open(os.path.join(BEST_MODEL_DIR, "metadata.txt"), "w") as f:
             f.write(f"Iteration: {iteration_num}\n")
             f.write(f"Timestamp: {log_entry['Timestamp']}\n")
             f.write(f"Model: {model_used}\n")
             f.write(f"val_excess_return: {val_excess_return}\n")
             f.write(f"val_max_drawdown: {val_max_drawdown}\n")
             f.write(f"Hypothesis: {log_entry['Hypothesis_Summary']}\n")
    else:
         print("Hypothesis rejected. Reverting train.py to previous best state.")
         shutil.copyfile("train.py.best", "train.py")
         
    # Optional Discord Webhook Notification
    DISCORD_WEBHOOK_URL = os.getenv("DISCORD_WEBHOOK_URL")
    if DISCORD_WEBHOOK_URL:
        emoji = "🚀" if is_new_best else "❌"
        msg = f"{emoji} **Autoresearch Iteration {iteration_num}**\n**Status:** {status}\n**Alpha (Excess Return):** {val_excess_return}\n**Max Drawdown:** {val_max_drawdown}\n**Hypothesis:** {log_entry['Hypothesis_Summary']}"
        try:
            requests.post(DISCORD_WEBHOOK_URL, json={"content": msg}, timeout=10)
        except Exception as e:
            print(f"Failed to trigger Discord webhook: {e}")
            
    return True

def main():
    print("Starting Autonomous Quant Researcher Loop...")
    print(f"Model chain: {' -> '.join(OPENROUTER_MODELS)} -> gemini-direct/{GEMINI_DIRECT_MODEL}")

    iter_num = 1
    # Check what iteration we are on
    if os.path.exists(LOG_FILE):
        logs = pd.read_csv(LOG_FILE)
        if not logs.empty:
            iter_num = logs['Iteration'].max() + 1

    # Wall-clock budget: exit after 50 min so the CI job has time to commit before GitHub's 6h limit
    CI_TIME_BUDGET = 50 * 60
    loop_start = time.time()

    while True:
        try:
            run_experiment(iter_num)
            iter_num += 1
            elapsed = time.time() - loop_start
            if elapsed >= CI_TIME_BUDGET:
                print(f"CI time budget reached ({elapsed/60:.1f} min). Exiting to commit results.")
                break
            print("\nWaiting 10 seconds before next iteration...")
            time.sleep(10)
        except KeyboardInterrupt:
            print("\nAutonomous Loop Stopped by User.")
            break

if __name__ == '__main__':
    main()
